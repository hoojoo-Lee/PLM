from datetime import datetime
from typing import List
import json
import os
import re
import shutil
import time
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, status, UploadFile
from sqlalchemy.orm import Session

from database import get_db
import models
from schemas import (
    BOMItemBrief,
    BOMItemCreate,
    BOMItemResponse,
    BOMItemUpdate,
    BOMVersionCreate,
    BOMVersionDetailResponse,
    BOMVersionResponse,
    BOMVersionUpdate,
)

router = APIRouter(prefix="/products/{product_id}/bom-versions", tags=["BOM"])


def _extract_drive_id(input_str: str) -> str:
    if not input_str:
        return ''
    trimmed = input_str.strip()
    file_d_regex = re.compile(r'/file/d/([-\w]+)')
    open_id_regex = re.compile(r'/open\?id=([-\w]+)')
    query_id_regex = re.compile(r'[?&]id=([-\w]+)')
    match = file_d_regex.search(trimmed) or open_id_regex.search(trimmed) or query_id_regex.search(trimmed)
    if match:
        return match.group(1)
    long_match = re.search(r'[-\w]{25,}', trimmed)
    if long_match:
        return long_match.group(0)
    return trimmed


def _get_diff(old_obj, new_obj, fields):
    diff = []
    for field in fields:
        old_val = getattr(old_obj, field, None)
        new_val = getattr(new_obj, field, None)
        if old_val != new_val:
            diff.append({
                "field": field,
                "old_value": old_val,
                "new_value": new_val
            })
    return diff


# =============================================================================
# BOM 版本 API
# =============================================================================

@router.post("", response_model=BOMVersionResponse, status_code=status.HTTP_201_CREATED)
def create_bom_version(product_id: int, payload: BOMVersionCreate, db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    existing = db.query(models.BOMVersion).filter(
        models.BOMVersion.product_id == product_id,
        models.BOMVersion.bom_type == payload.bom_type,
        models.BOMVersion.variant_tag == payload.variant_tag
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="BOM type and variant tag combination already exists")
    
    bom_version = models.BOMVersion(
        product_id=product_id,
        version_code=payload.version_code,
        bom_type=payload.bom_type,
        variant_tag=payload.variant_tag,
        status=payload.status,
        type_specific_fields=payload.type_specific_fields,
        change_notes=payload.change_notes,
        created_by=payload.created_by,
        released_at=payload.released_at,
    )
    db.add(bom_version)
    db.commit()
    db.refresh(bom_version)
    return bom_version


@router.get("", response_model=List[BOMVersionResponse])
def list_bom_versions(
    product_id: int,
    bom_type: str | None = None,
    variant_tag: str | None = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    q = db.query(models.BOMVersion).filter(models.BOMVersion.product_id == product_id)
    if bom_type is not None:
        q = q.filter(models.BOMVersion.bom_type == bom_type)
    if variant_tag is not None:
        q = q.filter(models.BOMVersion.variant_tag == variant_tag)
    versions = q.offset(skip).limit(limit).all()
    return versions


@router.get("/{version_id}", response_model=BOMVersionDetailResponse)
def get_bom_version(product_id: int, version_id: int, db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    version = db.query(models.BOMVersion).filter(
        models.BOMVersion.id == version_id,
        models.BOMVersion.product_id == product_id
    ).first()
    if not version:
        raise HTTPException(status_code=404, detail="BOMVersion not found")
    return version


@router.put("/{version_id}", response_model=BOMVersionResponse)
def update_bom_version(product_id: int, version_id: int, payload: BOMVersionUpdate, db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    version = db.query(models.BOMVersion).filter(
        models.BOMVersion.id == version_id,
        models.BOMVersion.product_id == product_id
    ).first()
    if not version:
        raise HTTPException(status_code=404, detail="BOMVersion not found")
    
    if payload.bom_type is not None and payload.variant_tag is not None:
        existing = db.query(models.BOMVersion).filter(
            models.BOMVersion.product_id == product_id,
            models.BOMVersion.bom_type == payload.bom_type,
            models.BOMVersion.variant_tag == payload.variant_tag,
            models.BOMVersion.id != version_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="BOM type and variant tag combination already exists")
    
    if payload.version_code is not None:
        version.version_code = payload.version_code
    if payload.bom_type is not None:
        version.bom_type = payload.bom_type
    if payload.variant_tag is not None:
        version.variant_tag = payload.variant_tag
    if payload.status is not None:
        version.status = payload.status
    if payload.type_specific_fields is not None:
        version.type_specific_fields = payload.type_specific_fields
    if payload.change_notes is not None:
        version.change_notes = payload.change_notes
    if payload.created_by is not None:
        version.created_by = payload.created_by
    if payload.released_at is not None:
        version.released_at = payload.released_at
    
    db.add(version)
    db.commit()
    db.refresh(version)
    return version


@router.delete("/{version_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_bom_version(product_id: int, version_id: int, db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    version = db.query(models.BOMVersion).filter(
        models.BOMVersion.id == version_id,
        models.BOMVersion.product_id == product_id
    ).first()
    if not version:
        raise HTTPException(status_code=404, detail="BOMVersion not found")
    
    db.delete(version)
    db.commit()
    return None


@router.patch("/{version_id}/archive", response_model=BOMVersionResponse)
def archive_bom_version(product_id: int, version_id: int, db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    version = db.query(models.BOMVersion).filter(
        models.BOMVersion.id == version_id,
        models.BOMVersion.product_id == product_id
    ).first()
    if not version:
        raise HTTPException(status_code=404, detail="BOMVersion not found")
    if version.status == "archived":
        raise HTTPException(status_code=400, detail="BOMVersion is already archived")
    
    version.status = "archived"
    db.add(version)
    db.commit()
    db.refresh(version)
    return version


@router.patch("/{version_id}/release", response_model=BOMVersionResponse)
def release_bom_version(product_id: int, version_id: int, db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    version = db.query(models.BOMVersion).filter(
        models.BOMVersion.id == version_id,
        models.BOMVersion.product_id == product_id
    ).first()
    if not version:
        raise HTTPException(status_code=404, detail="BOMVersion not found")
    
    version.status = "released"
    version.released_at = datetime.utcnow()
    db.add(version)
    db.commit()
    db.refresh(version)
    return version


@router.get("/{version_id}/changelog", response_model=List[dict])
def get_bom_version_changelog(product_id: int, version_id: int, db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    version = db.query(models.BOMVersion).filter(
        models.BOMVersion.id == version_id,
        models.BOMVersion.product_id == product_id
    ).first()
    if not version:
        raise HTTPException(status_code=404, detail="BOMVersion not found")
    
    item_ids = [row[0] for row in db.query(models.BOMItem.id).filter(
        models.BOMItem.bom_version_id == version_id
    ).all()]
    
    if not item_ids:
        return []
    
    changelogs = db.query(models.ChangeLog).filter(
        models.ChangeLog.entity_type == "BOMItem",
        models.ChangeLog.entity_id.in_(item_ids)
    ).order_by(models.ChangeLog.created_at.desc()).limit(200).all()
    
    items = db.query(models.BOMItem).filter(
        models.BOMItem.id.in_(item_ids)
    ).all()
    item_map = {item.id: {"name": item.name, "mpn": item.mpn} for item in items}
    
    result = []
    for cl in changelogs:
        info = item_map.get(cl.entity_id, {"name": "未知", "mpn": "-"})
        result.append({
            "id": cl.id,
            "created_at": cl.created_at.isoformat() if cl.created_at else None,
            "change_type": cl.change_type,
            "change_summary": cl.change_summary,
            "change_detail": cl.change_detail,
            "item_name": info["name"],
            "item_mpn": info["mpn"],
            "item_id": cl.entity_id,
        })
    return result


# =============================================================================
# BOM 物料项 API
# =============================================================================

@router.post("/{version_id}/items", response_model=BOMItemResponse, status_code=status.HTTP_201_CREATED)
def create_bom_item(product_id: int, version_id: int, payload: BOMItemCreate, db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    bom_version = db.query(models.BOMVersion).filter(
        models.BOMVersion.id == version_id,
        models.BOMVersion.product_id == product_id
    ).first()
    if not bom_version:
        raise HTTPException(status_code=404, detail="BOMVersion not found")
    
    bom_item = models.BOMItem(
        bom_version_id=version_id,
        category=payload.category,
        responsible_party=payload.responsible_party or 'NexPCB',
        mpn=payload.mpn,
        internal_pn=payload.internal_pn,
        name=payload.name,
        quantity=payload.quantity,
        responsible=payload.responsible,
        design_finalization=payload.design_finalization,
        picture_drive_id=payload.picture_drive_id,
        design_files=payload.design_files or [],
        item_specific_data=payload.item_specific_data or {},
        sample_no=payload.sample_no,
        qc_sample=payload.qc_sample,
        standard=payload.standard,
        comments=payload.comments,
        lead_time_s1=payload.lead_time_s1,
        lead_time_s2=payload.lead_time_s2,
        lead_time_s3=payload.lead_time_s3,
        sort_order=payload.sort_order,
    )
    db.add(bom_item)
    db.commit()
    db.refresh(bom_item)
    
    changelog = models.ChangeLog(
        entity_type="BOMItem",
        entity_id=bom_item.id,
        entity_name=bom_item.name,
        change_type="create",
        source="manual",
        change_summary=f"新增物料: {bom_item.name}",
        change_detail=[{"action": "create", "mpn": bom_item.mpn}],
    )
    db.add(changelog)
    db.commit()
    
    return bom_item


@router.get("/{version_id}/items", response_model=List[BOMItemResponse])
def list_bom_items(
    product_id: int,
    version_id: int,
    category: str | None = None,
    responsible_party: str | None = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    q = db.query(models.BOMItem).filter(models.BOMItem.bom_version_id == version_id)
    if category is not None:
        q = q.filter(models.BOMItem.category == category)
    if responsible_party is not None:
        q = q.filter(models.BOMItem.responsible_party == responsible_party)
    items = q.order_by(models.BOMItem.sort_order.asc()).offset(skip).limit(limit).all()
    return items


@router.put("/{version_id}/items/reorder", response_model=dict)
def reorder_bom_items(product_id: int, version_id: int, payload: list[dict], db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    for item_data in payload:
        item = db.query(models.BOMItem).filter(
            models.BOMItem.id == item_data.get("id"),
            models.BOMItem.bom_version_id == version_id
        ).first()
        if item:
            item.sort_order = item_data.get("sort_order", 0)
    db.commit()
    return {"status": "ok"}


@router.get("/{version_id}/items/{item_id}", response_model=BOMItemResponse)
def get_bom_item(product_id: int, version_id: int, item_id: int, db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    item = db.query(models.BOMItem).filter(
        models.BOMItem.id == item_id,
        models.BOMItem.bom_version_id == version_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOMItem not found")
    return item


@router.put("/{version_id}/items/{item_id}", response_model=BOMItemResponse)
def update_bom_item(product_id: int, version_id: int, item_id: int, payload: BOMItemUpdate, db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    item = db.query(models.BOMItem).filter(
        models.BOMItem.id == item_id,
        models.BOMItem.bom_version_id == version_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOMItem not found")
    
    fields_to_compare = [
        "category", "responsible_party", "mpn", "internal_pn", "name", "quantity",
        "responsible", "design_finalization", "picture_drive_id", "design_files",
        "item_specific_data", "sample_no", "qc_sample", "standard", "comments",
        "lead_time_s1", "lead_time_s2", "lead_time_s3", "sort_order"
    ]
    
    old_values = {f: getattr(item, f) for f in fields_to_compare}
    
    if payload.category is not None:
        item.category = payload.category
    if payload.responsible_party is not None:
        item.responsible_party = payload.responsible_party
    if payload.mpn is not None:
        item.mpn = payload.mpn
    if payload.internal_pn is not None:
        item.internal_pn = payload.internal_pn
    if payload.name is not None:
        item.name = payload.name
    if payload.quantity is not None:
        item.quantity = payload.quantity
    if payload.responsible is not None:
        item.responsible = payload.responsible
    if payload.design_finalization is not None:
        item.design_finalization = payload.design_finalization
    if payload.picture_drive_id is not None:
        item.picture_drive_id = _extract_drive_id(payload.picture_drive_id)
    if payload.design_files is not None:
        item.design_files = payload.design_files
    if payload.item_specific_data is not None:
        item.item_specific_data = payload.item_specific_data
    if payload.sample_no is not None:
        item.sample_no = payload.sample_no
    if payload.qc_sample is not None:
        item.qc_sample = payload.qc_sample
    if payload.standard is not None:
        item.standard = payload.standard
    if payload.comments is not None:
        item.comments = payload.comments
    if payload.lead_time_s1 is not None:
        item.lead_time_s1 = payload.lead_time_s1
    if payload.lead_time_s2 is not None:
        item.lead_time_s2 = payload.lead_time_s2
    if payload.lead_time_s3 is not None:
        item.lead_time_s3 = payload.lead_time_s3
    if payload.sort_order is not None:
        item.sort_order = payload.sort_order
    
    new_values = {f: getattr(item, f) for f in fields_to_compare}
    diff = []
    for field in fields_to_compare:
        if old_values[field] != new_values[field]:
            diff.append({
                "field": field,
                "old_value": old_values[field],
                "new_value": new_values[field]
            })
    
    if diff:
        changelog = models.ChangeLog(
            entity_type="BOMItem",
            entity_id=item.id,
            entity_name=item.name,
            change_type="update",
            source="manual",
            change_summary=f"更新物料: {item.name}",
            change_detail=diff,
        )
        db.add(changelog)
    
    db.add(item)
    db.commit()
    db.refresh(item)
    
    return item


@router.delete("/{version_id}/items/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_bom_item(product_id: int, version_id: int, item_id: int, db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    item = db.query(models.BOMItem).filter(
        models.BOMItem.id == item_id,
        models.BOMItem.bom_version_id == version_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOMItem not found")
    
    changelog = models.ChangeLog(
        entity_type="BOMItem",
        entity_id=item.id,
        entity_name=item.name,
        change_type="delete",
        source="manual",
        change_summary=f"删除物料: {item.name}",
        change_detail=[{"action": "delete", "mpn": item.mpn}],
    )
    db.add(changelog)
    db.delete(item)
    db.commit()
    return None


@router.post("/{version_id}/items/batch", response_model=List[BOMItemResponse], status_code=status.HTTP_201_CREATED)
def create_bom_items_batch(product_id: int, version_id: int, items: List[BOMItemCreate], db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    bom_version = db.query(models.BOMVersion).filter(
        models.BOMVersion.id == version_id,
        models.BOMVersion.product_id == product_id
    ).first()
    if not bom_version:
        raise HTTPException(status_code=404, detail="BOMVersion not found")
    
    created_items = []
    for payload in items:
        bom_item = models.BOMItem(
            bom_version_id=version_id,
            category=payload.category,
            responsible_party=payload.responsible_party or 'NexPCB',
            mpn=payload.mpn,
            internal_pn=payload.internal_pn,
            name=payload.name,
            quantity=payload.quantity,
            responsible=payload.responsible,
            design_finalization=payload.design_finalization,
            picture_drive_id=payload.picture_drive_id,
            design_files=payload.design_files or [],
            item_specific_data=payload.item_specific_data or {},
            sample_no=payload.sample_no,
            qc_sample=payload.qc_sample,
            standard=payload.standard,
            comments=payload.comments,
            lead_time_s1=payload.lead_time_s1,
            lead_time_s2=payload.lead_time_s2,
            lead_time_s3=payload.lead_time_s3,
            sort_order=payload.sort_order,
        )
        db.add(bom_item)
        created_items.append(bom_item)
    
    db.commit()
    for item in created_items:
        db.refresh(item)
    
    return created_items


@router.post("/{version_id}/import/parse", response_model=dict)
def parse_bom_excel(product_id: int, version_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    bom_version = db.query(models.BOMVersion).filter(
        models.BOMVersion.id == version_id,
        models.BOMVersion.product_id == product_id
    ).first()
    if not bom_version:
        raise HTTPException(status_code=404, detail="BOMVersion not found")
    
    allowed_extensions = ['.xlsx', '.xls', '.csv']
    filename = file.filename.lower() if file.filename else ''
    if not any(filename.endswith(ext) for ext in allowed_extensions):
        raise HTTPException(status_code=400, detail="Only Excel (.xlsx, .xls) or CSV files are allowed")
    
    try:
        if filename.endswith('.csv'):
            import csv
            content = file.file.read().decode('utf-8').strip()
            lines = content.split('\n')
            reader = csv.DictReader(lines)
            rows = list(reader)
        else:
            from io import BytesIO
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(file.file.read()), data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue
                rows.append(dict(zip(headers, row)))
        
        column_mapping = {
            'mpn': ['mpn', '料号', '物料编号', 'part number', 'pn', '编号'],
            'name': ['name', '物料名称', '名称', '品名', 'description', '描述'],
            'quantity': ['quantity', '数量', '用量', 'qty'],
            'category': ['category', '分类', '类别', '类型'],
            'responsible_party': ['responsible_party', '责任方', '模块', '板卡'],
            'responsible': ['responsible', '负责人', '责任人'],
        }
        
        def find_column(header, candidates):
            if not header:
                return None
            h = str(header).strip().lower()
            for candidate in candidates:
                if candidate.lower() in h or h in candidate.lower():
                    return h
            return None
        
        col_indices = {}
        if rows:
            first_row = rows[0]
            for key, candidates in column_mapping.items():
                for header in first_row.keys():
                    if find_column(header, candidates):
                        col_indices[key] = header
                        break
        
        parsed_items = []
        for row in rows:
            mpn = str(row.get(col_indices.get('mpn'))).strip() if col_indices.get('mpn') else ''
            name = str(row.get(col_indices.get('name'))).strip() if col_indices.get('name') else ''
            
            if not mpn and not name:
                continue
            
            quantity = row.get(col_indices.get('quantity'))
            try:
                quantity = int(quantity) if quantity else 1
            except ValueError:
                quantity = 1
            
            category = str(row.get(col_indices.get('category'))).strip() if col_indices.get('category') else ''
            responsible_party = str(row.get(col_indices.get('responsible_party'))).strip() if col_indices.get('responsible_party') else ''
            responsible = str(row.get(col_indices.get('responsible'))).strip() if col_indices.get('responsible') else ''
            
            if not name:
                name = mpn
            
            parsed_items.append({
                'category': category,
                'responsible_party': responsible_party,
                'mpn': mpn,
                'name': name,
                'quantity': quantity,
                'responsible': responsible,
            })
        
        return {
            'message': '解析成功',
            'parsed_items': parsed_items,
            'column_mapping': col_indices,
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"解析失败: {str(e)}")


@router.post("/{version_id}/import/commit", response_model=dict)
def commit_bom_import(product_id: int, version_id: int, items: List[BOMItemCreate], db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    bom_version = db.query(models.BOMVersion).filter(
        models.BOMVersion.id == version_id,
        models.BOMVersion.product_id == product_id
    ).first()
    if not bom_version:
        raise HTTPException(status_code=404, detail="BOMVersion not found")
    
    created_count = 0
    for payload in items:
        bom_item = models.BOMItem(
            bom_version_id=version_id,
            category=payload.category,
            responsible_party=payload.responsible_party or 'NexPCB',
            mpn=payload.mpn,
            internal_pn=payload.internal_pn,
            name=payload.name,
            quantity=payload.quantity,
            responsible=payload.responsible,
            design_finalization=payload.design_finalization,
            picture_drive_id=payload.picture_drive_id,
            design_files=payload.design_files or [],
            item_specific_data=payload.item_specific_data or {},
            sample_no=payload.sample_no,
            qc_sample=payload.qc_sample,
            standard=payload.standard,
            comments=payload.comments,
            lead_time_s1=payload.lead_time_s1,
            lead_time_s2=payload.lead_time_s2,
            lead_time_s3=payload.lead_time_s3,
            sort_order=payload.sort_order,
        )
        db.add(bom_item)
        created_count += 1
    
    db.commit()
    
    changelog = models.ChangeLog(
        entity_type="BOMVersion",
        entity_id=version_id,
        entity_name=bom_version.version_code,
        change_type="update",
        source="excel_import",
        change_summary=f"Excel导入: 新增 {created_count} 条物料",
        change_detail=[{"action": "excel_import", "count": created_count}],
    )
    db.add(changelog)
    db.commit()
    
    return {
        'message': '导入成功',
        'created_count': created_count,
    }


@router.post("/{version_id}/items/{item_id}/upload-picture", response_model=BOMItemResponse)
def upload_bom_item_picture(product_id: int, version_id: int, item_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    product = db.query(models.Product).get(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    item = db.query(models.BOMItem).filter(
        models.BOMItem.id == item_id,
        models.BOMItem.bom_version_id == version_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="BOMItem not found")
    
    try:
        upload_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")
        os.makedirs(upload_dir, exist_ok=True)
        
        ext = os.path.splitext(file.filename or "image.png")[1] or ".png"
        filename = f"{int(time.time())}_{item_id}{ext}"
        file_path = os.path.join(upload_dir, filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        old_picture_id = item.picture_drive_id
        relative_path = f"/uploads/{filename}"
        item.picture_drive_id = relative_path
        
        db.add(item)
        db.commit()
        db.refresh(item)
        
        changelog = models.ChangeLog(
            entity_type="BOMItem",
            entity_id=item.id,
            entity_name=item.name,
            change_type="update",
            source="manual",
            change_summary=f"更新预览图: {item.name}",
            change_detail=[{"action": "update_picture", "old_picture_id": old_picture_id, "new_picture_id": relative_path}],
        )
        db.add(changelog)
        db.commit()
        
        return item
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"图片上传失败: {str(e)}")